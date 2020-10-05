import os
import sys
import glob
import pandas as pd
import openpyxl as px
import tkinter as tk
import tkinter.filedialog as fl
import tkinter.messagebox as mb
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import PatternFill


class SelectFile:
    def set_file():
        filetype = [("all file", "*.csv")]
        iDir = glob.glob(r"C:\Users\めっき実験系列\Desktop\測温データ\**\**\**\*")
        new_file = max(iDir, key=os.path.getmtime)
        file_path = fl.askopenfilename(initialdir=new_file, filetypes=filetype)
        file1.set(file_path)
        

class MakeExcel:
    def main():
        excel_name = str(os.path.splitext(file1.get())[0]) + "-1.xlsx"
        
        try:
            df = pd.read_csv(file1.get(), skiprows=57,
                             usecols=[2], encoding="cp932")
        except FileNotFoundError:
            mb.showinfo("", "ファイルが選択されていません")
            sys.exit()
            
        df.drop(df.tail(3).index, inplace=True)
        df_float = df.astype("float").round(1)

        # 同名ファイルがある場合、上書き保存するか確認
        if os.path.isfile(excel_name):
            res = mb.askquestion("", "同名ファイルがあります。上書きしますか？")
            if res == "yes":
                df_float.to_excel(excel_name, header=False, index=False)
            elif res == "no":
                mb.showinfo("", "もう一度ファイル名を確認してください")
                sys.exit()
        else:
            df_float.to_excel(excel_name, header=False, index=False)

        wb = px.load_workbook(excel_name)
        ws = wb.active
        sheet = wb["Sheet1"]
        sc = sheet.cell
        wc = ws.cell

        sheet.insert_cols(0, 1)  # 昇温時間の列を追加

        start = 1
        cell_diff = 0

        # 3行下の温度と比べ10以上上昇した場合昇温開始とする
        while cell_diff <= 10:
            cell_diff = float(sc(row=start+3, column=2).value)\
                        - float(sc(row=start, column=2).value)
            start += 1

        end = start
        v1 = 0
        fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    
        # 昇温時間を0.5ずつ入力
        while sc(row=end, column=2).value is not None:
            wc(row=end, column=1, value=v1)
            end += 1
            v1 += 0.5
            try:
                if int(entry_time.get()) == v1:
                    for col in range(1, 4):
                        wc(row=end, column=col).fill = fill
                        max_entry_time = end
            except ValueError:
                pass

        try:
            temp_var = int(entry_temp.get()) - 10  # 狙い温度-10℃
        except ValueError:
            mb.showinfo("", "狙い温度が入力されていません")
            sys.exit()

        keep = 1
    
        # 狙い温度-10℃の行
        while sc(row=keep, column=2).value <= temp_var:
            keep += 1

        # 小数点第一位が5の場合、1行下げる
        if str(sc(row=keep, column=1).value)[-1] == str(5):
            keep = keep + 1
        
        # 保定開始時間のセルに色を付ける
        for col in range(1, 3):
            wc(row=keep, column=col).fill = fill

        v2 = 0
    
        # 狙い温度-10℃から、保定時間を0.5ずつ入力
        while keep != end:
            wc(row=keep, column=3, value=v2)
            keep += 1
            v2 += 0.5
            # 該当の保定時間に色を付ける
            for time in entry_list:
                if int(time.get()) == v2:
                    for col in range(1, 4):
                        wc(row=keep, column=col).fill = fill
                        max_entry_time = keep

        # セルの書式を小数第一位で揃える
        for row in sheet:
            for cell in row:
                cell.number_format = "0.0"

        chart = ScatterChart()
    
        x_values = Reference(ws, min_row=start, min_col=1,
                             max_row=end, max_col=1)  # 時間
        y_values = Reference(ws, min_row=start, min_col=2,
                             max_row=end, max_col=2)  # 温度

        graph = Series(y_values, x_values)
        chart.series.append(graph)

        try:
            ws.add_chart(chart, "D"+str(max_entry_time))  # チャートを保定終了時の行に表示
        except UnboundLocalError:
            mb.showinfo("", "時間が入力されていません")
            sys.exit()
    
        wb.save(excel_name)
        mb.showinfo("", "Excelファイルを作成しました")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("CSVをExcelに変換")

    # frame1
    frame1 = tk.LabelFrame(root, text="ファイルを選択")
    frame1.grid(row=0, columnspan=2, sticky="we", padx=5)

    select_button = tk.Button(frame1, text="選択",
                              command=SelectFile.set_file, width=10)
    select_button.grid(row=0, column=3)

    # ファイルパスの表示
    file1 = tk.StringVar()
    file1_entry = tk.Entry(frame1, textvariable=file1, width=35)
    file1_entry.grid(row=0, column=2, padx=5)

    # frame2
    frame2 = tk.LabelFrame(root, text="条件")
    frame2.grid(row=1, sticky="we")

    text_temp = tk.Label(frame2, text="狙い温度（℃）", width=20)
    text_temp.grid(row=0, column=0, padx=5)

    entry_temp = tk.Entry(frame2, width=15)
    entry_temp.grid(row=1, column=0, padx=5)

    text_time = tk.Label(frame2, text="在炉時間（秒）", width=20)
    text_time.grid(row=2, column=0, padx=5)

    entry_time = tk.Entry(frame2, width=15)
    entry_time.grid(row=3, column=0, padx=5)

    text_keep = tk.Label(frame2, text="保定時間（秒）:複数指定可", width=25)
    text_keep.grid(row=0, column=1)

    entry_list = []

    for n in range(1, 4):
        entry = tk.Entry(frame2, width=15)
        entry.grid(row=n, column=1, padx=5, pady=5)
        entry.insert(tk.END, 0)
        entry_list.append(entry)

    action_button = tk.Button(frame2, text="実行",
                              command=MakeExcel.main, width=15)
    action_button.grid(row=4, column=0)

    root.mainloop()
